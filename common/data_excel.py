__author__ = "luo"

import openpyxl
import os


def get_data_all(excel_name, sheet_name):
    r"""
    通过excel文件名以及表名，去读取对应文件，获得数据

    :param excel_name: excel的名称
    :param sheet_name: 表名
    :return 返回所有数据，列表形式
    """
    # 获取当前__file__文件(执行文件)路径
    current_path = os.path.abspath(__file__)
    path_list = current_path.split("common")
    # 获取excel文件路径
    excel_path = os.path.join(path_list[0], "testData", excel_name + ".xlsx")
    # 打开excel
    wb = openpyxl.load_workbook(excel_path)
    # 获取所有工作表
    sheet_names = wb.sheetnames
    # 遍历所有工作表，获得需要的表
    data_sheet = None
    for i in range(len(sheet_names)):
        if sheet_names[i] == sheet_name:
            # 打开此工作表(对象)
            data_sheet = wb[sheet_names[i]]
            break
        else:
            # 表不存在时，抛出异常
            raise Exception(u"没有在文件%s.xlsx中找到%s表" % (excel_name, sheet_name))
    # 获取第一行数据（标题行）
    list_data_row_1 = []
    for row in data_sheet[1]:
        list_data_row_1.append(row.value)

    # 读取所有数据（除标题行）
    list_data_all = []
    for row in data_sheet.iter_rows(min_row=2):
        i = 0
        row_dict = dict()
        for cell in row:
            if list_data_row_1[i] == "params" \
                or list_data_row_1[i] == "headers" \
                    or list_data_row_1[i] == "body":  # 处理json数据
                if cell.value:
                    row_dict[list_data_row_1[i]] = eval(cell.value.replace("\n", "").replace(" ", ""))
            else:
                row_dict[list_data_row_1[i]] = cell.value
            i += 1
        list_data_all.append(row_dict)
    wb.close()
    return list_data_all


def get_data_by_key(data_all, url_key):
    """
    根据url_key返回数据

    :param data_all: 全部数据，列表形式
    :param url_key: key
    :return: 返回列表
    """
    list_data_key = []
    for i in range(len(data_all)):
        if data_all[i]["url_key"] == url_key:
            list_data_key.append(data_all[i])
    return list_data_key


